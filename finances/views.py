import datetime
import decimal
import hashlib
import operator

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate
from django.contrib.auth.models import Group, Permission
from django.contrib.messages.views import SuccessMessageMixin
from django.core.exceptions import ObjectDoesNotExist, PermissionDenied
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Q
from django.http import Http404
from django.shortcuts import HttpResponse, redirect, render
from django.urls import reverse
from django.utils.encoding import force_text
from django.utils.timezone import now
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import FormView, View
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from borgia.utils import (GroupLateralMenuFormMixin, GroupLateralMenuMixin,
                          GroupPermissionMixin, UserMixin, shop_from_group)
from finances.forms import (BankAccountCreateForm, BankAccountUpdateForm,
                            ExceptionnalMovementForm,
                            GenericListSearchDateForm, RechargingListForm,
                            SaleListSearchDateForm, SelfLydiaCreateForm,
                            SelfTransfertCreateForm, SharedEventAddWeightForm,
                            SharedEventCreateForm, SharedEventDeleteForm,
                            SharedEventDownloadXlsxForm, SharedEventFinishForm,
                            SharedEventListForm,
                            SharedEventSelfRegistrationForm,
                            SharedEventUpdateForm, SharedEventUploadXlsxForm,
                            UserSupplyMoneyForm)
from finances.models import (BankAccount, Cash, Cheque, ExceptionnalMovement,
                             LydiaFaceToFace, LydiaOnline, Recharging, Sale,
                             SharedEvent, Transfert)
from notifications.models import notify
from settings_data.models import Setting
from settings_data.utils import settings_safe_get
from users.models import User


class UserBankAccountCreate(GroupPermissionMixin, UserMixin, FormView,
                            GroupLateralMenuFormMixin):
    """
    View to create a bank account for a specific user.

    :param kwargs['group_name']: name of the group, mandatory
    :param kwargs['user_pk']: pk of the specific user, mandatory
    :type kwargs['user_pk']: positif integer
    :type kwargs['group_name']: string
    :raises: Http404 if the group_name doesn't match a group
    :raises: Http404 if the user_pk doesn't match an user
    :raises: Http404 if the group doesn't have the permission to add account
    """
    template_name = 'finances/user_bankaccount_create.html'
    perm_codename = 'add_bankaccount'
    lm_active = None
    form_class = BankAccountCreateForm

    def form_valid(self, form):
        """
        Create a bank account for the user.
        """
        bank_account = BankAccount.objects.create(
            bank=form.cleaned_data['bank'],
            account=form.cleaned_data['account'],
            owner=self.user
        )
        # We notify
        notify(notification_class_name='user_bank_account_creation',
               actor=self.request.user,
               recipient=bank_account.owner,
               target_object=bank_account
               )
        return super(UserBankAccountCreate, self).form_valid(form)


class SelfBankAccountCreate(GroupPermissionMixin, FormView,
                            GroupLateralMenuFormMixin):
    """
    View to create a bank account for the logged user.

    :param kwargs['group_name']: name of the group, mandatory
    :type kwargs['group_name']: string
    :raises: Http404 if the group_name doesn't match a group
    :raises: Http404 if the logged user is not in the group
    """
    template_name = 'finances/self_bankaccount_create.html'
    perm_codename = None
    lm_active = None
    form_class = BankAccountCreateForm

    def form_valid(self, form):
        """
        Create a bank account for the logged user.
        """
        BankAccount.objects.create(
            bank=form.cleaned_data['bank'],
            account=form.cleaned_data['account'],
            owner=self.request.user
        )
        self.success_url = reverse(
            'url_self_user_update',
            kwargs={'group_name': self.group.name}
        )
        return super(SelfBankAccountCreate, self).form_valid(form)


class UserBankAccountUpdate(GroupPermissionMixin, UserMixin, FormView,
                            GroupLateralMenuFormMixin):
    """
    View to update a bank account for a specific user.

    :param kwargs['group_name']: name of the group, mandatory
    :param kwargs['user_pk']: pk of the specific user, mandatory
    :param kwargs['pk']: pk of the bankaccount, mandatory
    :type kwargs['user_pk']: positif integer
    :type kwargs['group_name']: string
    :type kwargs['pk']: positiv integer
    :raises: Http404 if the group_name doesn't match a group
    :raises: Http404 if the user_pk doesn't match an user
    :raises: Http404 if the group doesn't have the permission to update account
    :raises: Http404 if the pk doesn't match a bank account
    :raises: PermissionDenied if the bank account is not owned by the user
    """
    template_name = 'finances/user_bankaccount_update.html'
    perm_codename = 'change_bankaccount'
    lm_active = None
    form_class = BankAccountUpdateForm

    def dispatch(self, request, *args, **kwargs):
        """
        Check if the bank account is owned by the user.

        :raises: PermissionDenied if not.
        """
        try:
            self.object = BankAccount.objects.get(pk=kwargs['pk'])
            self.user = User.objects.get(pk=kwargs['user_pk'])
        except ObjectDoesNotExist:
            raise Http404
        if self.object.owner != self.user:
            raise PermissionDenied
        return super(UserBankAccountUpdate, self).dispatch(
            request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(UserBankAccountUpdate, self).get_context_data(**kwargs)
        context['object'] = self.object
        return context

    def get_initial(self):
        initial = super(UserBankAccountUpdate, self).get_initial()
        initial['bank'] = self.object.bank
        initial['account'] = self.object.account
        return initial

    def form_valid(self, form):
        self.object.bank = form.cleaned_data['bank']
        self.object.account = form.cleaned_data['account']
        self.object.save()
        # We notify
        notify(notification_class_name='user_bank_account_update',
               actor=self.request.user,
               recipient=self.object.owner,
               target_object=self.object
               )
        return super(UserBankAccountUpdate, self).form_valid(form)


class SelfBankAccountUpdate(GroupPermissionMixin, FormView,
                            GroupLateralMenuFormMixin):
    """
    View to update a bank account for the logged user.

    :param kwargs['group_name']: name of the group, mandatory
    :param kwargs['pk']: pk of the bankaccount, mandatory
    :type kwargs['group_name']: string
    :type kwargs['pk']: positiv integer
    :raises: Http404 if the group_name doesn't match a group
    :raises: Http404 if the pk doesn't match a bank account
    :raises: PermissionDenied if the bank account is not owned by the user
    """
    template_name = 'finances/self_bankaccount_update.html'
    perm_codename = None
    lm_active = None
    form_class = BankAccountUpdateForm

    def dispatch(self, request, *args, **kwargs):
        """
        Check if the bank account is owned by the logged user.

        :raises: PermissionDenied if not.
        """
        try:
            self.object = BankAccount.objects.get(pk=kwargs['pk'])
        except ObjectDoesNotExist:
            raise Http404
        if self.object.owner != request.user:
            raise PermissionDenied
        return super(SelfBankAccountUpdate, self).dispatch(
            request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(SelfBankAccountUpdate, self).get_context_data(**kwargs)
        context['object'] = self.object
        return context

    def get_initial(self):
        initial = super(SelfBankAccountUpdate, self).get_initial()
        initial['bank'] = self.object.bank
        initial['account'] = self.object.account
        return initial

    def form_valid(self, form):
        self.object.bank = form.cleaned_data['bank']
        self.object.account = form.cleaned_data['account']
        self.object.save()
        self.success_url = reverse(
            'url_self_user_update',
            kwargs={'group_name': self.group.name}
        )
        return super(SelfBankAccountUpdate, self).form_valid(form)


class UserBankAccountDelete(GroupPermissionMixin, UserMixin, View,
                            GroupLateralMenuMixin):
    """
    View to delete a bank account for a specific user.

    :param kwargs['group_name']: name of the group, mandatory
    :param kwargs['user_pk']: pk of the specific user, mandatory
    :param kwargs['pk']: pk of the bankaccount, mandatory
    :type kwargs['user_pk']: positif integer
    :type kwargs['group_name']: string
    :type kwargs['pk']: positiv integer
    :raises: Http404 if the group_name doesn't match a group
    :raises: Http404 if the user_pk doesn't match an user
    :raises: Http404 if the group doesn't have the permission to delete account
    :raises: Http404 if the pk doesn't match a bank account
    :raises: PermissionDenied if the bank account is not owned by the user
    """
    template_name = 'finances/user_bankaccount_delete.html'
    perm_codename = 'delete_bankaccount'
    lm_active = None

    def dispatch(self, request, *args, **kwargs):
        """
        Check if the bank account is owned by the user.

        :raises: PermissionDenied if not.
        """
        try:
            self.object = BankAccount.objects.get(pk=kwargs['pk'])
            self.user = User.objects.get(pk=kwargs['user_pk'])
        except ObjectDoesNotExist:
            raise Http404
        if self.object.owner != self.user:
            raise PermissionDenied
        return super(UserBankAccountDelete, self).dispatch(
            request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        context['object'] = self.object
        return render(request, self.template_name, context=context)

    def post(self, request, *args, **kwargs):
        self.object.delete()
        return redirect(force_text(self.success_url))


class SelfBankAccountDelete(GroupPermissionMixin, View, GroupLateralMenuMixin):
    """
    View to delete a bank account for the logged user.

    :param kwargs['group_name']: name of the group, mandatory
    :param kwargs['pk']: pk of the bankaccount, mandatory
    :type kwargs['group_name']: string
    :type kwargs['pk']: positiv integer
    :raises: Http404 if the group_name doesn't match a group
    :raises: Http404 if the pk doesn't match a bank account
    :raises: PermissionDenied if the bank account is not owned by the logged
    user
    """
    template_name = 'finances/self_bankaccount_delete.html'
    perm_codename = None
    lm_active = None

    def dispatch(self, request, *args, **kwargs):
        """
        Check if the bank account is owned by the logged user.

        :raises: PermissionDenied if not.
        """
        try:
            self.object = BankAccount.objects.get(pk=kwargs['pk'])
        except ObjectDoesNotExist:
            raise Http404
        if self.object.owner != request.user:
            raise PermissionDenied
        return super(SelfBankAccountDelete, self).dispatch(
            request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        context['object'] = self.object
        return render(request, self.template_name, context=context)

    def post(self, request, *args, **kwargs):
        self.object.delete()
        return redirect(reverse(
            'url_self_user_update',
            kwargs={'group_name': self.group.name}))


class SaleList(GroupPermissionMixin, FormView, GroupLateralMenuFormMixin):
    """
    View to list sales.

    If the group derived from a shop, only sales from this shop are listed.
    Else (presidents, treasurers or vice-presidents-internal) all sales are
    listed with this view. The view handle request of the form to filter
    sales.
    :note:: only sales are listed here. Sales come from a shop, for other
    types of transactions, please refer to other classes (RechargingList,
    TransfertList and ExceptionnalMovementList).

    :param kwargs['group_name']: name of the group, mandatory
    :type kwargs['group_name']: string
    :raises: Http404 if the group_name doesn't match a group
    :raises: Http404 if the group doesn't have the permission to list sales
    """
    template_name = 'finances/sale_shop_list.html'
    perm_codename = 'list_sale'
    lm_active = 'lm_sale_list'
    form_class = GenericListSearchDateForm

    query_shop = None
    search = None
    date_begin = None
    date_end = None

    def dispatch(self, request, *args, **kwargs):
        try:
            self.group = Group.objects.get(name=kwargs['group_name'])
        except ObjectDoesNotExist:
            raise Http404

        try:
            self.shop = shop_from_group(self.group)
        except ValueError:
            self.template_name = 'finances/sale_list.html'
            self.form_class = SaleListSearchDateForm

        return super(SaleList, self).dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(SaleList, self).get_context_data(**kwargs)
        sales_tab_header = []
        seen = set(sales_tab_header)
        page = self.request.POST.get('page', 1)

        try:
            context['sale_list'] = Sale.objects.filter(
                shop=self.shop).order_by('-datetime')
        except AttributeError:
            context['sale_list'] = Sale.objects.all().order_by('-datetime')

        # The sale_list is paginated by passing the filtered QuerySet to Paginator
        paginator = Paginator(self.form_query(context['sale_list']), 50)
        try:
            # The requested page is grabbed
            sales = paginator.page(page)
        except PageNotAnInteger:
            # If the requested page is not an integer
            sales = paginator.page(1)
        except EmptyPage:
            # If the requested page is out of range, the last page is grabbed
            sales = paginator.page(paginator.num_pages)

        context['sale_list'] = sales

        for sale in context['sale_list']:
            if sale.from_shop() not in seen:
                seen.add(sale.from_shop())
                sales_tab_header.append(sale.from_shop())

        context['sales_tab_header'] = sales_tab_header

        return context

    def form_query(self, query):
        if self.search:
            query = query.filter(
                Q(operator__last_name__icontains=self.search)
                | Q(operator__first_name__icontains=self.search)
                | Q(operator__surname__icontains=self.search)
                | Q(operator__username__icontains=self.search)
                | Q(recipient__last_name__icontains=self.search)
                | Q(recipient__first_name__icontains=self.search)
                | Q(recipient__surname__icontains=self.search)
                | Q(recipient__username__icontains=self.search)
                | Q(sender__last_name__icontains=self.search)
                | Q(sender__first_name__icontains=self.search)
                | Q(sender__surname__icontains=self.search)
                | Q(sender__username__icontains=self.search)
            )

        if self.date_begin:
            query = query.filter(
                datetime__gte=self.date_begin)

        if self.date_end:
            query = query.filter(
                datetime__lte=self.date_end)

        if self.query_shop:
            query = query.filter(shop=self.query_shop)

        return query

    def form_valid(self, form):
        if form.cleaned_data['search']:
            self.search = form.cleaned_data['search']

        if form.cleaned_data['date_begin']:
            self.date_begin = form.cleaned_data['date_begin']

        if form.cleaned_data['date_end']:
            self.date_end = form.cleaned_data['date_end']
        try:
            if form.cleaned_data['shop']:
                self.query_shop = form.cleaned_data['shop']
        except KeyError:
            pass

        return self.get(self.request, self.args, self.kwargs)


class SaleRetrieve(GroupPermissionMixin, View, GroupLateralMenuMixin):
    """
    Retrieve a sale.

    A sale comes from a shop, for other type of transaction, please refer to
    other classes (RechargingRetrieve, TransfertRetrieve,
    ExceptionnalMovementRetrieve).

    :param kwargs['group_name']: name of the group, mandatory
    :param kwargs['pk']: pk of the sale, mandatory
    :type kwargs['group_name']: string
    :type kwargs['pk']: positiv integer
    :raises: Http404 if group_name doesn't match a group
    :raises: Http404 if the pk doesn't match a sale
    :raises: Http404 if the sale is not from a shop
    :raises: Http404 if the sale is not from the shop derived from the group
    :raises: Http404 if the group doesn't have the permission to retrieve sale
    """
    template_name = 'finances/sale_retrieve.html'
    perm_codename = 'retrieve_sale'

    def dispatch(self, request, *args, **kwargs):
        try:
            self.object = Sale.objects.get(pk=kwargs['pk'])
        except ObjectDoesNotExist:
            raise Http404

        try:
            self.group = Group.objects.get(name=kwargs['group_name'])
        except ObjectDoesNotExist:
            raise Http404

        if self.object.from_shop() is None:
            raise Http404

        try:
            self.shop = shop_from_group(self.group)
            if self.object.from_shop() != self.shop:
                raise Http404
        except ValueError:
            pass

        return super(SaleRetrieve, self).dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        context['object'] = self.object
        return render(request, self.template_name, context=context)


class RechargingList(GroupPermissionMixin, FormView,
                     GroupLateralMenuFormMixin):
    """
    View to list recharging sales.

    The view handle request of the form to filter recharging sales.
    :note:: only recharging sales are listed here. For other types of
    transactions, please refer to other classes (SaleList, TransfertList and
    ExceptionnalMovementList).

    :param kwargs['group_name']: name of the group, mandatory
    :type kwargs['group_name']: string
    :raises: Http404 if the group_name doesn't match a group
    :raises: Http404 if the group doesn't have the permission to list
    recharging sales
    """
    template_name = 'finances/recharging_list.html'
    perm_codename = 'list_recharging'
    lm_active = 'lm_recharging_list'
    form_class = RechargingListForm

    search = None
    date_end = now() + datetime.timedelta(days=1)
    date_begin = now() - datetime.timedelta(days=7)
    operators = None

    def get_context_data(self, **kwargs):
        context = super(RechargingList, self).get_context_data(**kwargs)

        context['recharging_list'] = Recharging.objects.all().order_by(
            '-datetime')

        context['recharging_list'] = self.form_query(
            context['recharging_list'])[:1000]

        context['info'] = self.info(context['recharging_list'])
        return context

    def get_initial(self, **kwargs):
        initial = super(RechargingList, self).get_initial(**kwargs)
        initial['date_begin'] = self.date_begin
        initial['date_end'] = self.date_end
        return initial

    def info(self, query):
        info = {
            'cash': {
                'total': 0,
                'nb': 0,
                'ids': Cash.objects.none()
            },
            'cheque': {
                'total': 0,
                'nb': 0,
                'ids': Cheque.objects.none()
            },
            'lydia_face2face': {
                'total': 0,
                'nb': 0,
                'ids': LydiaFaceToFace.objects.none()
            },
            'lydia_online': {
                'total': 0,
                'nb': 0,
                'ids': LydiaOnline.objects.none()
            },
            'total': {
                'total': 0,
                'nb': 0
            }
        }

        for r in query:
            if r.payment_solution.get_type() == 'cash':
                info['cash']['total'] += r.payment_solution.amount
                info['cash']['nb'] += 1
                info['cash']['ids'] |= Cash.objects.filter(
                    pk=r.payment_solution.cash.pk)
            if r.payment_solution.get_type() == 'cheque':
                info['cheque']['total'] += r.payment_solution.amount
                info['cheque']['nb'] += 1
                info['cheque']['ids'] |= Cheque.objects.filter(
                    pk=r.payment_solution.cheque.pk)
            if r.payment_solution.get_type() == 'lydiafacetoface':
                info['lydia_face2face']['total'] += r.payment_solution.amount
                info['lydia_face2face']['nb'] += 1
                info['lydia_face2face']['ids'] |= LydiaFaceToFace.objects.filter(
                    pk=r.payment_solution.lydiafacetoface.pk)
            if r.payment_solution.get_type() == 'lydiaonline':
                info['lydia_online']['total'] += r.payment_solution.amount
                info['lydia_online']['nb'] += 1
                info['lydia_online']['ids'] |= LydiaOnline.objects.filter(
                    pk=r.payment_solution.lydiaonline.pk)
            info['total']['total'] += r.payment_solution.amount
            info['total']['nb'] += 1
        return info

    def form_query(self, query):
        if self.search:
            query = query.filter(
                Q(operator__last_name__contains=self.search)
                | Q(operator__first_name__contains=self.search)
                | Q(operator__surname__contains=self.search)
                | Q(sender__last_name__contains=self.search)
                | Q(sender__first_name__contains=self.search)
                | Q(sender__surname__contains=self.search)
            )

        if self.date_begin:
            query = query.filter(
                datetime__gte=self.date_begin)

        if self.date_end:
            query = query.filter(
                datetime__lte=self.date_end)

        if self.operators:
            query = query.filter(operator__in=self.operators)

        return query

    def form_valid(self, form):
        if form.cleaned_data['search'] != '':
            self.search = form.cleaned_data['search']

        if form.cleaned_data['date_begin'] != '':
            self.date_begin = form.cleaned_data['date_begin']

        if form.cleaned_data['date_end'] != '':
            self.date_end = form.cleaned_data['date_end']

        if form.cleaned_data['operators']:
            self.operators = form.cleaned_data['operators']

        context = self.get_context_data()
        return self.get(self.request, self.args, self.kwargs)


class RechargingRetrieve(GroupPermissionMixin, View, GroupLateralMenuMixin):
    """
    Retrieve a recharging sale.

    For other type of transaction, please refer to other classes (SaleRetrieve,
    TransfertRetrieve, ExceptionnalMovementRetrieve).

    :param kwargs['group_name']: name of the group, mandatory
    :param kwargs['pk']: pk of the sale, mandatory
    :type kwargs['group_name']: string
    :type kwargs['pk']: positiv integer
    :raises: Http404 if group_name doesn't match a group
    :raises: Http404 if the pk doesn't match a sale
    :raises: Http404 if the sale is not a recharging sale
    :raises: Http404 if the group doesn't have the permission to retrieve
    recharging sale
    """
    template_name = 'finances/recharging_retrieve.html'
    perm_codename = 'retrieve_recharging'

    def dispatch(self, request, *args, **kwargs):
        try:
            self.object = Recharging.objects.get(pk=kwargs['pk'])
        except ObjectDoesNotExist:
            raise Http404

        return super(RechargingRetrieve, self).dispatch(request, *args,
                                                        **kwargs)

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        context['object'] = self.object
        return render(request, self.template_name, context=context)


class TransfertList(GroupPermissionMixin, FormView, GroupLateralMenuFormMixin):
    """
    View to list transfert sales.

    The view handle request of the form to filter transfert sales.
    :note:: only transfert sales are listed here. For other types of
    transactions, please refer to other classes (SaleList, RechargingList and
    ExceptionnalMovementList).

    :param kwargs['group_name']: name of the group, mandatory
    :type kwargs['group_name']: string
    :raises: Http404 if the group_name doesn't match a group
    :raises: Http404 if the group doesn't have the permission to list transfert
    sales
    """
    template_name = 'finances/transfert_list.html'
    perm_codename = 'list_transfert'
    lm_active = 'lm_transfert_list'
    form_class = GenericListSearchDateForm

    search = None
    date_begin = None
    date_end = None

    def get_context_data(self, **kwargs):
        context = super(TransfertList, self).get_context_data(**kwargs)

        context['transfert_list'] = Transfert.objects.all().order_by('-datetime')

        context['transfert_list'] = self.form_query(
            context['transfert_list'])[:100]

        return context

    def form_query(self, query):
        if self.search:
            query = query.filter(
                Q(recipient__last_name__contains=self.search)
                | Q(recipient__first_name__contains=self.search)
                | Q(recipient__surname__contains=self.search)
                | Q(sender__last_name__contains=self.search)
                | Q(sender__first_name__contains=self.search)
                | Q(sender__surname__contains=self.search)
            )

        if self.date_begin:
            query = query.filter(
                datetime__gte=self.date_begin)

        if self.date_end:
            query = query.filter(
                datetime__lte=self.date_end)

        return query

    def form_valid(self, form):
        if form.cleaned_data['search'] != '':
            self.search = form.cleaned_data['search']

        if form.cleaned_data['date_begin'] != '':
            self.date_begin = form.cleaned_data['date_begin']

        if form.cleaned_data['date_end'] != '':
            self.date_end = form.cleaned_data['date_end']

        context = self.get_context_data()
        return self.get(self.request, self.args, self.kwargs)


class TransfertRetrieve(GroupPermissionMixin, View, GroupLateralMenuMixin):
    """
    Retrieve a transfert sale.

    For other type of transaction, please refer to other classes (SaleRetrieve,
    RechargingRetrieve, ExceptionnalMovementRetrieve).

    :param kwargs['group_name']: name of the group, mandatory
    :param kwargs['pk']: pk of the sale, mandatory
    :type kwargs['group_name']: string
    :type kwargs['pk']: positiv integer
    :raises: Http404 if group_name doesn't match a group
    :raises: Http404 if the pk doesn't match a sale
    :raises: Http404 if the sale is not a transfert sale
    :raises: Http404 if the group doesn't have the permission to retrieve
    transfert sale
    """
    template_name = 'finances/transfert_retrieve.html'
    perm_codename = 'retrieve_transfert'

    def dispatch(self, request, *args, **kwargs):
        try:
            self.object = Transfert.objects.get(pk=kwargs['pk'])
        except ObjectDoesNotExist:
            raise Http404

        return super(TransfertRetrieve, self).dispatch(request, *args,
                                                       **kwargs)

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        context['object'] = self.object
        return render(request, self.template_name, context=context)


class SelfTransfertCreate(GroupPermissionMixin, SuccessMessageMixin, FormView,
                          GroupLateralMenuFormMixin):
    template_name = 'finances/self_transfert_create.html'
    perm_codename = 'add_transfert'
    lm_active = 'lm_self_transfert_create'
    form_class = SelfTransfertCreateForm
    success_message = "Le montant de %(amount)s€ a bien été transféré à %(recipient)s."

    def get_form_kwargs(self, **kwargs):
        kwargs = super(SelfTransfertCreate, self).get_form_kwargs(**kwargs)
        kwargs['sender'] = self.request.user
        return kwargs

    def form_valid(self, form):
        # Get user or ValidationError
        recipient = form.cleaned_data['recipient']

        transfert = Transfert.objects.create(
            sender=self.request.user,
            recipient=recipient,
            amount=form.cleaned_data['amount'],
            justification=form.cleaned_data['justification']
        )
        transfert.pay()
        # We notify
        notify(notification_class_name='transfer_creation',
               actor=self.request.user,
               recipient=recipient,
               target_object=transfert)
        return super(SelfTransfertCreate, self).form_valid(form)

    def get_success_message(self, cleaned_data):
        return self.success_message % dict(
            amount=cleaned_data['amount'],
            recipient=cleaned_data['recipient']
        )


class ExceptionnalMovementList(GroupPermissionMixin, FormView,
                               GroupLateralMenuFormMixin):
    """
    View to list exceptionnal movement sales.

    The view handle request of the form to filter recharging exceptionnal
    movement.
    :note:: only exceptionnal movement sales are listed here. For other types
    of transactions, please refer to other classes (SaleList, TransfertList and
    SaleList).

    :param kwargs['group_name']: name of the group, mandatory
    :type kwargs['group_name']: string
    :raises: Http404 if the group_name doesn't match a group
    :raises: Http404 if the group doesn't have the permission to list
    exceptionnal movement sales
    """
    template_name = 'finances/exceptionnalmovement_list.html'
    perm_codename = 'list_exceptionnal_movement'
    lm_active = 'lm_exceptionnalmovement_list'
    form_class = GenericListSearchDateForm

    search = None
    date_begin = None
    date_end = None

    def get_context_data(self, **kwargs):
        context = super(ExceptionnalMovementList, self).get_context_data(
            **kwargs)

        context['exceptionnalmovement_list'] = ExceptionnalMovement.objects.all(
        ).order_by('-datetime')

        context['exceptionnalmovement_list'] = self.form_query(
            context['exceptionnalmovement_list'])[:100]

        return context

    def form_query(self, query):
        if self.search:
            query = query.filter(
                Q(operator__last_name__contains=self.search)
                | Q(operator__first_name__contains=self.search)
                | Q(operator__surname__contains=self.search)
                | Q(recipient__last_name__contains=self.search)
                | Q(recipient__first_name__contains=self.search)
                | Q(recipient__surname__contains=self.search)
            )

        if self.date_begin:
            query = query.filter(
                datetime__gte=self.date_begin)

        if self.date_end:
            query = query.filter(
                datetime__lte=self.date_end)

        return query

    def form_valid(self, form):
        if form.cleaned_data['search'] != '':
            self.search = form.cleaned_data['search']

        if form.cleaned_data['date_begin'] != '':
            self.date_begin = form.cleaned_data['date_begin']

        if form.cleaned_data['date_end'] != '':
            self.date_end = form.cleaned_data['date_end']

        context = self.get_context_data()
        return self.get(self.request, self.args, self.kwargs)


class ExceptionnalMovementRetrieve(GroupPermissionMixin, View,
                                   GroupLateralMenuMixin):
    """
    Retrieve an exceptionnal movement sale.

    For other type of transaction, please refer to other classes (SaleRetrieve,
    TransfertRetrieve, RechargingRetrieve).

    :param kwargs['group_name']: name of the group, mandatory
    :param kwargs['pk']: pk of the sale, mandatory
    :type kwargs['group_name']: string
    :type kwargs['pk']: positiv integer
    :raises: Http404 if group_name doesn't match a group
    :raises: Http404 if the pk doesn't match a sale
    :raises: Http404 if the sale is not an exceptionnal movement sale
    :raises: Http404 if the group doesn't have the permission to retrieve
    exceptionnal movement sale
    """
    template_name = 'finances/exceptionnalmovement_retrieve.html'
    perm_codename = 'retrieve_exceptionnal_movement'

    def dispatch(self, request, *args, **kwargs):
        try:
            self.object = ExceptionnalMovement.objects.get(pk=kwargs['pk'])
        except ObjectDoesNotExist:
            raise Http404

        return super(ExceptionnalMovementRetrieve, self).dispatch(
            request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        context['object'] = self.object
        return render(request, self.template_name, context=context)


class SelfTransactionList(GroupPermissionMixin, FormView,
                          GroupLateralMenuFormMixin):
    """
    View to list transactions of the logged user.

    :param kwargs['group_name']: name of the group, mandatory
    :type kwargs['group_name']: string
    :raises: Http404 if the group_name doesn't match a group
    """
    template_name = 'finances/self_transaction_list.html'
    perm_codename = None
    lm_active = 'lm_self_transaction_list'
    form_class = GenericListSearchDateForm

    search = None
    date_begin = None
    date_end = None

    def get_context_data(self, **kwargs):
        context = super(SelfTransactionList, self).get_context_data(**kwargs)

        context['transaction_list'] = self.form_query(
            self.request.user.list_transaction())[:100]
        return context

    # TODO: form to be used
    def form_query(self, query):
        return query

    def form_valid(self, form):
        if form.cleaned_data['search']:
            self.search = form.cleaned_data['search']
        if form.cleaned_data['date_begin']:
            self.date_begin = form.cleaned_data['date_begin']

        if form.cleaned_data['date_end']:
            self.date_end = form.cleaned_data['date_end']
        try:
            if form.cleaned_data['shop']:
                self.query_shop = form.cleaned_data['shop']
        except KeyError:
            pass
        context = self.get_context_data()
        return self.get(self.request, self.args, self.kwargs)


class UserExceptionnalMovementCreate(GroupPermissionMixin, UserMixin, FormView,
                                     GroupLateralMenuFormMixin):
    """
    View to create an exceptionnal movement (debit or credit) for a specific
    user.

    :param kwargs['group_name']: name of the group, mandatory
    :param kwargs['user_pk']: pk of the specific user, mandatory
    :type kwargs['user_pk']: positif integer
    :type kwargs['group_name']: string
    :raises: Http404 if the group_name doesn't match a group
    :raises: Http404 if the user_pk doesn't match an user
    :raises: Http404 if the group doesn't have the permission to add an
    exceptionnal movement
    """
    template_name = 'finances/user_exceptionnalmovement_create.html'
    perm_codename = 'add_exceptionnal_movement'
    lm_active = None
    form_class = ExceptionnalMovementForm

    def form_valid(self, form):
        """
        :note:: The form is assumed clean, then the couple username/password
        for the operator is right.
        """
        operator = authenticate(
            username=form.cleaned_data['operator_username'],
            password=form.cleaned_data['operator_password'])
        amount = form.cleaned_data['amount']
        is_credit = False

        if form.cleaned_data['type_movement'] == 'credit':
            is_credit = True

        exceptionnal_movement = ExceptionnalMovement.objects.create(
            justification=form.cleaned_data['justification'],
            operator=operator,
            recipient=self.user,
            is_credit=is_credit,
            amount=amount
        )
        exceptionnal_movement.pay()

        return super(UserExceptionnalMovementCreate, self).form_valid(form)

    def get_initial(self):
        """
        Populate the form with the current login user for the operator (only
        username of course).
        """
        initial = super(UserExceptionnalMovementCreate, self).get_initial()
        initial['operator_username'] = self.request.user.username
        return initial


class UserSupplyMoney(GroupPermissionMixin, UserMixin, FormView,
                      GroupLateralMenuFormMixin):
    template_name = 'finances/user_supplymoney.html'
    perm_codename = 'supply_money_user'
    lm_active = None
    form_class = UserSupplyMoneyForm

    def get_form_kwargs(self, **kwargs):
        kwargs = super(UserSupplyMoney, self).get_form_kwargs(**kwargs)
        kwargs['user'] = self.user
        return kwargs

    def form_valid(self, form):
        """
        :note:: The form is assumed clean, then the couple username/password
        for the operator is right.
        """
        operator = authenticate(
            username=form.cleaned_data['operator_username'],
            password=form.cleaned_data['operator_password'])
        sender = self.user
        payment = []
        if form.cleaned_data['type'] == 'cheque':
            cheque = Cheque.objects.create(
                amount=form.cleaned_data['amount'],
                signature_date=form.cleaned_data['signature_date'],
                cheque_number=form.cleaned_data['unique_number'],
                sender=sender,
                recipient=User.objects.get(username='AE_ENSAM'),
                bank_account=form.cleaned_data['bank_account'])
            payment = cheque

        elif form.cleaned_data['type'] == 'cash':
            cash = Cash.objects.create(
                sender=sender,
                recipient=User.objects.get(username='AE_ENSAM'),
                amount=form.cleaned_data['amount'])
            payment = cash

        elif form.cleaned_data['type'] == 'lydia':
            lydia = LydiaFaceToFace.objects.create(
                date_operation=form.cleaned_data['signature_date'],
                id_from_lydia=form.cleaned_data['unique_number'],
                sender=sender,
                recipient=User.objects.get(username='AE_ENSAM'),
                amount=form.cleaned_data['amount'])
            payment = lydia

        recharging = Recharging.objects.create(
            sender=sender,
            operator=operator,
            payment_solution=payment
        )
        recharging.pay()

        return super(UserSupplyMoney, self).form_valid(form)

    def get_initial(self):
        """
        Populate the form with the current login user for the operator (only
        username of course).
        """
        initial = super(UserSupplyMoney, self).get_initial()
        initial['signature_date'] = now
        initial['operator_username'] = self.request.user.username
        return initial


class SelfLydiaCreate(GroupPermissionMixin, FormView,
                      GroupLateralMenuFormMixin):
    """
    View to supply himself by Lydia.

    :param kwargs['group_name']: name of the group, mandatory
    :type kwargs['group_name']: string
    :raises: Http404 if the group_name doesn't match a group
    """
    form_class = SelfLydiaCreateForm
    template_name = 'finances/self_lydia_create.html'
    perm_codename = None
    lm_active = 'lm_self_lydia_create'

    def get_form_kwargs(self):
        kwargs = super(SelfLydiaCreate, self).get_form_kwargs()

        # Min value is always 0.01
        try:
            min_value = Setting.objects.get(
                name='LYDIA_MIN_PRICE').get_value()
            if min_value is not None:
                if min_value > 0:
                    kwargs['min_value'] = decimal.Decimal(min_value)
                else:
                    kwargs['min_value'] = decimal.Decimal("0.01")
            else:
                kwargs['min_value'] = decimal.Decimal("0.01")
        except ObjectDoesNotExist:
            kwargs['min_value'] = decimal.Decimal("0.01")

        try:
            max_value = Setting.objects.get(
                name='LYDIA_MAX_PRICE').get_value()
            if max_value is not None:
                kwargs['max_value'] = decimal.Decimal(max_value)
            else:
                kwargs['max_value'] = None
        except ObjectDoesNotExist:
            kwargs['max_value'] = None

        return kwargs

    def form_valid(self, form):
        """
        Save the current phone number as default phone number for the user.
        Render the Lydia button.
        """
        user = self.request.user
        if user.phone is None:
            user.phone = form.cleaned_data['tel_number']
            user.save()

        context = super(SelfLydiaCreate, self).get_context_data()
        context['vendor_token'] = settings_safe_get(
            "LYDIA_VENDOR_TOKEN").get_value()
        context['confirm_url'] = settings.LYDIA_CONFIRM_URL
        context['callback_url'] = settings.LYDIA_CALLBACK_URL
        context['amount'] = form.cleaned_data['amount']
        context['tel_number'] = form.cleaned_data['tel_number']
        context['message'] = (
            "Borgia - AE ENSAM - Crédit de "
            + user.__str__())
        return render(self.request,
                      'finances/self_lydia_button.html',
                      context=context)

    def get_initial(self):
        initial = super(SelfLydiaCreate, self).get_initial()
        initial['tel_number'] = self.request.user.phone
        return initial

    def get_context_data(self, **kwargs):
        context = super(SelfLydiaCreate, self).get_context_data(**kwargs)
        try:
            if settings_safe_get("LYDIA_API_TOKEN").get_value() in ['', 'non définie']:
                context['no_module'] = True
            else:
                context['no_module'] = False
        except:
            context['no_module'] = True
        return context


class SelfLydiaConfirm(GroupPermissionMixin, View, GroupLateralMenuMixin):
    """
    View to confirm supply by Lydia.

    This view is only here to have some indications for the user. It does
    nothing more and can be "generated" with GET irelevant GET parameters.

    :note:: transaction and order parameters are given by Lydia but aren't
    used here.

    :param kwargs['group_name']: name of the group, mandatory
    :type kwargs['group_name']: string
    :param GET['transaction']: id of the Lydia transaction, mandatory.
    :type GET['transaction']: string
    :param GET['order']: id of the order from Lydia, mandatory.
    :type GET['order']: string
    :raises: Http404 if the group_name doesn't match a group
    """
    perm_codename = None
    template_name = 'finances/self_lydia_confirm.html'

    # TODO: check if a Lydia object exist and if it's from the current day,
    # else raise Error
    def get(self, request, *args, **kwargs):
        context = super(SelfLydiaConfirm, self).get_context_data()
        context['transaction'] = self.request.GET.get('transaction')
        context['order'] = self.request.GET.get('order_ref')
        return render(request, self.template_name, context=context)


class SharedEventList(GroupPermissionMixin, FormView,
                      GroupLateralMenuFormMixin):
    template_name = 'finances/sharedevent_list.html'
    lm_active = 'lm_sharedevent_list'
    perm_codename = 'list_sharedevent'
    form_class = SharedEventListForm

    def dispatch(self, request, *args, **kwargs):
        try:
            self.group = Group.objects.get(name=kwargs['group_name'])
        except ObjectDoesNotExist:
            raise Http404
        # For safety, but it shouldn't be possible with regex pattern in url FOR NOW (If Post is used, url need to be modified)
        except ValueError:
            raise Http404

        return super(SharedEventList, self).dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(SharedEventList, self).get_context_data(**kwargs)
        shared_events = SharedEvent.objects.filter(
            date__gte = datetime.date.today(), done=False).order_by('-date')

        for se in shared_events:  # Duplicate
            # Si fini, on recupere la participation, sinon la preinscription
            se.weight_of_user = se.get_weight_of_user(
                self.request.user, se.done)
            se.number_registrants = se.get_number_registrants()
            se.number_participants = se.get_number_participants()
            se.total_weights_registrants = se.get_total_weights_registrants()
            se.total_weights_participants = se.get_total_weights_participants()
            se.has_perm_manage = (self.request.user == se.manager or
                                  Permission.objects.get(codename='manage_sharedevent') in self.group.permissions.all())
        context['shared_events'] = shared_events
        # Permission SelfRegistration
        if Permission.objects.get(codename='self_register_sharedevent') in self.group.permissions.all():
            context['has_perm_self_register_sharedevent'] = True

        return context

    def form_valid(self, form, **kwargs):

        date_begin = form.cleaned_data['date_begin']
        date_end = form.cleaned_data['date_end']
        order_by = form.cleaned_data['order_by']
        done = form.cleaned_data['done']

        if date_begin is None:
            date_begin = datetime.date.min
        if date_end is None:
            date_end = datetime.date.max

        if done != "both":
            if done == "done":
                done = True
            else:
                done = False

            shared_events = SharedEvent.objects.filter(
                date__range=[date_begin, date_end], done=done)
        else:
            shared_events = SharedEvent.objects.filter(
                date__range=[date_begin, date_end])

        context = self.get_context_data(**kwargs)
        if order_by != '-date':
            shared_events = shared_events.order_by(order_by).order_by('-date')
        else:
            shared_events = shared_events.order_by('-date')

        for se in shared_events:  # Duplicate
            # Si fini, on recupere la participation, sinon la preinscription
            se.weight_of_user = se.get_weight_of_user(
                self.request.user, se.done)
            se.number_registrants = se.get_number_registrants()
            se.number_participants = se.get_number_participants()
            se.total_weights_registrants = se.get_total_weights_registrants()
            se.total_weights_participants = se.get_total_weights_participants()

        context['shared_events'] = shared_events

        return self.render_to_response(context)


class SharedEventCreate(GroupPermissionMixin, SuccessMessageMixin, FormView,
                        GroupLateralMenuFormMixin):
    form_class = SharedEventCreateForm
    template_name = 'finances/sharedevent_create.html'
    success_url = None
    perm_codename = 'add_sharedevent'
    lm_active = 'lm_sharedevent_create'
    success_message = "'%(description)s' a bien été créé."

    def form_valid(self, form):

        se = SharedEvent.objects.create(
            description=form.cleaned_data['description'],
            date=form.cleaned_data['date'],
            allow_self_registeration=form.cleaned_data['allow_self_registeration'],
            manager=self.request.user)
        if form.cleaned_data['price']:
            se.price = decimal.Decimal(form.cleaned_data['price'])
        if form.cleaned_data['bills']:
            se.bills = form.cleaned_data['bills']
        if form.cleaned_data['date_end_registration']:
            date_end_registration = form.cleaned_data['date_end_registration']
            if date_end_registration > datetime.date.today() and date_end_registration <= form.cleaned_data['date']:
                se.date_end_registration = form.cleaned_data['date_end_registration']

        se.save()
        self.se = se

        return super(SharedEventCreate, self).form_valid(form)

    def get_success_message(self, cleaned_data):
        return self.success_message % dict(
            description=self.se.description,
        )

    def get_success_url(self):
        return reverse('url_sharedevent_update',
                       kwargs={'group_name': self.group.name, 'pk': self.se.pk}
                       )


class SharedEventDelete(GroupPermissionMixin, SuccessMessageMixin, FormView, GroupLateralMenuMixin):
    """
    Delete a sharedevent and redirect to the list of sharedevents.

    :param kwargs['group_name']: name of the group used.
    :param kwargs['pk']: pk of the event
    :param self.perm_codename: codename of the permission checked.
    """
    template_name = 'finances/sharedevent_delete.html'
    form_class = SharedEventDeleteForm
    success_url = None
    perm_codename = None  # Checked in dispatch
    success_message = "L'évènement a bien été supprimé."

    def dispatch(self, request, *args, **kwargs):
        try:
            self.group = Group.objects.get(name=kwargs['group_name'])
            self.se = SharedEvent.objects.get(pk=kwargs['pk'])
        except ObjectDoesNotExist:
            raise Http404

        # Permissions
        try:
            if Permission.objects.get(codename='manage_sharedevent') not in self.group.permissions.all():
                if request.user != self.se.manager:
                    raise Http404
        except ObjectDoesNotExist:
            raise Http404
        if self.se.done:
            raise PermissionDenied

        return super(SharedEventDelete, self).dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(SharedEventDelete, self).get_context_data(**kwargs)
        context['se'] = self.se
        return context

    def form_valid(self, form):
        self.se.delete()
        return super(SharedEventDelete, self).form_valid(form)

    def get_success_url(self):
        return reverse('url_sharedevent_list',
                       kwargs={'group_name': self.group.name}
                       )


class SharedEventFinish(GroupPermissionMixin, SuccessMessageMixin, FormView, GroupLateralMenuFormMixin):
    """
    Finish a sharedevent and redirect to the list of sharedevents.
    This command is used when you want to keep the event in the database, but
    you don't want to pay in Borgia (for instance paid with real money).

    :param kwargs['group_name']: name of the group used.
    :param kwargs['pk']: pk of the event
    :param self.perm_codename: codename of the permission checked.
    """
    form_class = SharedEventFinishForm
    template_name = 'finances/sharedevent_finish.html'
    success_url = None
    perm_codename = None  # Checked in dispatch
    success_message = "'%(description)s' a bien été terminé."

    def dispatch(self, request, *args, **kwargs):
        try:
            self.se = SharedEvent.objects.get(pk=kwargs['pk'])
        except ObjectDoesNotExist:
            raise Http404

        # Permissions
        try:
            group = Group.objects.get(name=kwargs['group_name'])
            if Permission.objects.get(codename='proceed_payment_sharedevent') not in group.permissions.all():
                raise Http404
        except ObjectDoesNotExist:
            raise Http404

        if self.se.done:
            raise PermissionDenied

        # Check if there are participants
        self.total_weights_participants = self.se.get_total_weights_participants()
        if self.total_weights_participants == 0:
            return redirect(reverse(
                'url_sharedevent_update',
                kwargs={'group_name': group.name, 'pk': self.se.pk}
            ) + '?no_participant=True')

        # Get some data :
        self.price = self.se.price
        try:
            self.ponderation_price = self.price / self.total_weights_participants
        except TypeError:
            self.ponderation_price = 0

        return super(SharedEventFinish, self).dispatch(request, *args, **kwargs)

    def get_initial(self):
        """
        Populate the form with the current price.
        """
        initial = super(SharedEventFinish, self).get_initial()
        initial['total_price'] = self.price
        initial['ponderation_price'] = round(self.ponderation_price, 2)
        return initial

    def get_context_data(self, **kwargs):
        context = super(SharedEventFinish, self).get_context_data(**kwargs)
        context['se'] = self.se
        context['total_price'] = self.se.price
        context['total_weights_participants'] = self.total_weights_participants
        context['ponderation_price'] = round(self.ponderation_price, 2)
        return context

    def form_valid(self, form):
        type_payment = form.cleaned_data['type_payment']
        if (type_payment == 'pay_by_total'):
            self.se.pay_by_total(self.request.user, User.objects.get(
                pk=1), form.cleaned_data['total_price'])

        if (type_payment == 'pay_by_ponderation'):
            self.se.pay_by_ponderation(self.request.user, User.objects.get(
                pk=1), form.cleaned_data['ponderation_price'])

        if (type_payment == 'no_payment'):
            self.se.end_without_payment(form.cleaned_data['remark'])

        return super(SharedEventFinish, self).form_valid(form)

    def get_success_message(self, cleaned_data):
        return self.success_message % dict(
            description=self.se.description,
        )

    def get_success_url(self):
        return reverse('url_sharedevent_list',
                       kwargs={'group_name': self.group.name}
                       )


class SharedEventUpdate(GroupPermissionMixin, SuccessMessageMixin, FormView, GroupLateralMenuMixin):
    """
    Update the Shared Event,
    """
    form_class = SharedEventUpdateForm
    template_name = 'finances/sharedevent_update.html'
    perm_codename = None  # Checked in dispatch()
    success_message = "'%(description)s' a bien été mis à jour."

    manager_changed = False

    def dispatch(self, request, *args, **kwargs):
        try:
            self.group = Group.objects.get(name=kwargs['group_name'])
            self.se = SharedEvent.objects.get(pk=int(kwargs['pk']))
        except ObjectDoesNotExist:
            raise Http404
        # For safety, but it shouldn't be possible with regex pattern in url FOR NOW (If Post is used, url need to be modified)
        except ValueError:
            raise Http404

        # Permissions
        try:
            if Permission.objects.get(codename='manage_sharedevent') not in self.group.permissions.all():
                if request.user != self.se.manager:
                    raise Http404
        except ObjectDoesNotExist:
            raise Http404

        self.no_participant = request.GET.get(
            'no_participant', False) == "True"  # "True" == "True"

        return super(SharedEventUpdate, self).dispatch(request, *args, **kwargs)

    def get_initial(self):
        initial = super(SharedEventUpdate, self).get_initial()
        initial['price'] = self.se.price
        initial['bills'] = self.se.bills
        initial['manager'] = self.se.manager.username
        initial['allow_self_registeration'] = self.se.allow_self_registeration
        return initial

    def get_context_data(self, **kwargs):

        # list_year() contains smth like [2011, 2015, ...]

        context = super(SharedEventUpdate, self).get_context_data(**kwargs)
        context['pk'] = self.se.pk
        context['done'] = self.se.done
        if self.se.done:
            context['remark'] = self.se.remark
            context['price'] = self.se.price
        # Pour les users
        context['number_registrants'] = self.se.get_number_registrants
        context['number_participants'] = self.se.get_number_participants
        context['total_weights_registrants'] = self.se.get_total_weights_registrants
        context['total_weights_participants'] = self.se.get_total_weights_participants
        context['no_participant'] = self.no_participant

        # Création des forms excel
        context['upload_xlsx_form'] = SharedEventUploadXlsxForm()
        context['download_xlsx_form'] = SharedEventDownloadXlsxForm()

        # Permission FinishEvent
        try:
            group = Group.objects.get(name=context['group_name'])
            if Permission.objects.get(codename='proceed_payment_sharedevent') in group.permissions.all():
                context['has_perm_proceed_payment'] = True
        except ObjectDoesNotExist:
            pass  # has_perm not True

        return context

    def form_valid(self, form):
        if form.cleaned_data['price']:
            self.se.price = form.cleaned_data['price']
        if form.cleaned_data['bills']:
            self.se.bills = form.cleaned_data['bills']
        if form.cleaned_data['manager']:
            form_manager = User.objects.get(
                username=form.cleaned_data['manager'])
            manage_permission = False
            if self.se.manager != form_manager:
                for group in form_manager.groups.all():
                    if Permission.objects.get(codename='add_sharedevent') in group.permissions.all():
                        manage_permission = True
                        break
                if manage_permission:
                    self.se.manager = form_manager
                    self.manager_changed = True
                else:
                    messages.warning(self.request,
                                     "%(user)s ne dispose pas de droits suffisants pour gérer l'évènement" % dict(
                                         user=form_manager))
        self.se.allow_self_registeration = form.cleaned_data['allow_self_registeration']
        self.se.save()

        return super(SharedEventUpdate, self).form_valid(form)

    def get_success_message(self, cleaned_data):
        if self.manager_changed:
            return "%(user)s gère désormais l'évènement" % dict(
                user=self.se.manager,
            )
        else:
            return self.success_message % dict(
                description=self.se.description,
            )

    def get_success_url(self):
        if self.manager_changed and Permission.objects.get(codename='manage_sharedevent') not in self.group.permissions.all():
            return reverse('url_group_workboard',
                           kwargs={'group_name': self.group.name}
                           )
        else:
            return reverse('url_sharedevent_update',
                           kwargs={'group_name': self.group.name,
                                   'pk': self.se.pk}
                           )


class SharedEventSelfRegistration(GroupPermissionMixin, SuccessMessageMixin, FormView, GroupLateralMenuMixin):
    """
    Allow a user to register himself

    :param kwargs['group_name']: name of the group used.
    :param kwargs['pk']: pk of the event
    :param self.perm_codename: codename of the permission checked.
    """
    form_class = SharedEventSelfRegistrationForm
    template_name = 'finances/sharedevent_self_registration.html'
    perm_codename = 'self_register_sharedevent'

    def dispatch(self, request, *args, **kwargs):
        try:
            self.se = SharedEvent.objects.get(pk=int(kwargs['pk']))
        except ObjectDoesNotExist:
            raise Http404
        # For safety, but it shouldn't be possible with regex pattern in url FOR NOW (If Post is used, url need to be modified)
        except ValueError:
            raise Http404

        # Permissions
        if self.se.done:
            raise PermissionDenied

        if not self.se.allow_self_registeration:
            raise PermissionDenied

        if self.se.date_end_registration:
            if datetime.date.today() > self.se.date_end_registration:
                raise PermissionDenied

        return super(SharedEventSelfRegistration, self).dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(SharedEventSelfRegistration,
                        self).get_context_data(**kwargs)
        context['shared_event'] = self.se
        context['registeration_of_user'] = self.se.get_weight_of_user(
            self.request.user, False)  # Duplicate
        return context

    def get_initial(self):
        initial = super(SharedEventSelfRegistration, self).get_initial()
        initial['weight'] = self.se.get_weight_of_user(
            self.request.user, False)  # Duplicate
        return initial

    def form_valid(self, form):
        self.new_weight = int(form.cleaned_data['weight'])
        self.se.change_weight(self.request.user, self.new_weight, False)
        return super(SharedEventSelfRegistration, self).form_valid(form)

    def get_success_message(self, cleaned_data):
        if self.new_weight > 0:
            return "Vous avez bien été inscrit"
        else:
            return "Vous avez bien été désinscrit"

    def get_success_url(self):
        return reverse('url_sharedevent_self_registration',
                       kwargs={'group_name': self.group.name, 'pk': self.se.pk}
                       )


class SharedEventChangeWeight(GroupPermissionMixin, View):
    perm_codename = None

    def get(self, request, *args, **kwargs):
        """
        Change la valeur de la pondération d'un participant user pour un événement
        Permissions :   Si événements terminé -> denied,
                        Si pas manager ou pas la perm 'finances.manage_sharedevent' -> denied
        :param pk: pk de l'événement
        :param user_pk: paramètre GET correspondant au pk de l'user
        :param pond_pk: paramètre GET correspondant à la nouvelle pondération
        :type pk, user_pk, pond_pk: int
        """
        response = 0

        try:
            # Variables d'entrées
            se = SharedEvent.objects.get(pk=kwargs['pk'])
            user = User.objects.get(pk=kwargs['participant_pk'])
            pond = int(request.GET['pond'])
            isParticipant = int(request.GET['isParticipant'])  # Boolean

            # Permission
            if request.user != se.manager and request.user.has_perm('finances.manage_sharedevent') is False:
                raise PermissionDenied
            # Même en ayant la permission, on ne modifie plus une event terminé
            elif se.done is True:
                raise PermissionDenied

            if pond > 0:
                if isParticipant in [0, 1]:
                    # Changement de la pondération
                    se.change_weight(user, pond, isParticipant)

                    # Réponse
                    response = 1

        except KeyError:
            pass
        except ObjectDoesNotExist:
            pass
        except ValueError:
            pass

        return HttpResponse(response)


class SharedEventManageUsers(GroupPermissionMixin, FormView, GroupLateralMenuMixin):
    """
    Manage the users. The get displays the list of concerned users. The post form add weight to one of them.
    """
    form_class = SharedEventAddWeightForm
    template_name = 'finances/sharedevent_manage_users.html'
    perm_codename = None  # Checked in dispatch()

    def get_list_weights(self, state):
        se = SharedEvent.objects.get(pk=self.kwargs['pk'])
        if state == 'users':
            return se.list_users_weight()
        elif state == 'participants':
            return se.list_participants_weight()
        elif state == 'registrants':
            return se.list_registrants_weight()

    def dispatch(self, request, *args, **kwargs):
        try:
            self.group = Group.objects.get(name=kwargs['group_name'])
            self.se = SharedEvent.objects.get(pk=kwargs['pk'])
        except ObjectDoesNotExist:
            raise Http404

        # Permission
        if request.user != self.se.manager or not request.user.has_perm('finances.manage_sharedevent'):
            raise Http404

        return super(SharedEventManageUsers, self).dispatch(request, *args, **kwargs)

    def get_initial(self):
        initial = super(SharedEventManageUsers, self).get_initial()
        if self.request.GET.get('state') is not None and self.request.GET.get('state') == "registrants":
            initial['state'] = 'registered'
        else:
            initial['state'] = 'participant'
        return initial

    def get_context_data(self, **kwargs):

        # DEFAULT OPTIONS FOR LISTING
        state = 'users'
        order_by = 'username'

        # If an option is provided
        if self.request.GET.get('state') is not None:
            if self.request.GET.get('state') in ['users', 'participants', 'registrants']:
                state = self.request.GET.get('state')
        # If an option is provided
        if self.request.GET.get('order_by') is not None:
            if self.request.GET.get('order_by') in ['username', 'last_name', 'surname', 'year']:
                order_by = self.request.GET.get('order_by')

        initial_list_users_form = {
            'state': state,
            'order_by': order_by,
        }

        # Création des forms
        list_users_form = SharedEventListUsersForm(
            initial=initial_list_users_form)

        context = super(SharedEventManageUsers,
                        self).get_context_data(**kwargs)
        context['pk'] = self.se.pk
        context['done'] = self.se.done
        context['price'] = self.se.price
        context['state'] = state
        context['order_by'] = order_by

        context['list_users_form'] = list_users_form
        context['list_weights'] = sorted(self.get_list_weights(
            state), key=lambda item: getattr(item[0], order_by))
        return context

    def form_valid(self, form):
        user = form.cleaned_data['user']
        weight = form.cleaned_data['weight']
        # True pour un participant
        isParticipant = form.cleaned_data['state'] == 'participant'

        self.se.add_weight(user, weight, isParticipant)
        return super(SharedEventManageUsers, self).form_valid(form)

    def get_success_url(self):
        return reverse('url_sharedevent_manage_users',
                       kwargs={'group_name': self.group.name, 'pk': self.se.pk})


class SharedEventRemoveUser(GroupPermissionMixin, View):
    perm_codename = None

    def get(self, request, *args, **kwargs):
        se = SharedEvent.objects.get(pk=kwargs['pk'])

        # Permission
        if se.done is True:
            raise PermissionDenied
        if not (request.user == se.manager or request.user.has_perm('finances.manage_sharedevent')):
            raise PermissionDenied

        try:
            state = request.GET['state']
            order_by = request.GET['order_by']
            user_pk = kwargs['user_pk']

            if state == "users":
                if user_pk == 'ALL':
                    for u in se.users.all():
                        se.remove_user(u)
                else:
                    se.remove_user(User.objects.get(pk=user_pk))

            elif state == "participants":
                if user_pk == 'ALL':
                    for u in se.users.all():
                        se.change_weight(u, 0, True)
                else:
                    se.change_weight(User.objects.get(pk=user_pk), 0, True)

            elif state == "registrants":
                if user_pk == 'ALL':
                    for u in se.users.all():
                        se.change_weight(u, 0, False)
                else:
                    se.change_weight(User.objects.get(pk=user_pk), 0, False)

            else:
                raise Http404
        except ObjectDoesNotExist:
            raise Http404

        return redirect(reverse(
            'url_sharedevent_manage_users',
            kwargs={'group_name': self.group.name, 'pk': se.pk}
        ) + "?state=" + state + "&order_by=" + order_by + "#table_users")


class SharedEventDownloadXlsx(GroupPermissionMixin, FormView, GroupLateralMenuMixin):
    """
    Download Excel.
    """
    form_class = SharedEventDownloadXlsxForm
    perm_codename = None  # Checked in dispatch()

    def dispatch(self, request, *args, **kwargs):
        try:
            self.group = Group.objects.get(name=kwargs['group_name'])
            self.se = SharedEvent.objects.get(pk=kwargs['pk'])
        except ObjectDoesNotExist:
            raise Http404

        # Permission
        if not (request.user == self.se.manager or request.user.has_perm('finances.manage_sharedevent')):
            raise PermissionDenied

        return super(SharedEventDownloadXlsx, self).dispatch(request, *args, **kwargs)

    def form_valid(self, form):

        wb = Workbook()
        # grab the active worksheet
        ws = wb.active
        ws.title = "Test"
        ws.append(['Username', 'Pondération',
                   'Infos (Non utilisées) ->', 'Nom Prénom', 'Bucque'])

        if form.cleaned_data['state'] == 'year':

            if form.cleaned_data['years']:
                # Contains the years selected
                list_year_result = form.cleaned_data['years']

                users = User.objects.filter(year__in=list_year_result, is_active=True).exclude(
                    groups=Group.objects.get(pk=1)).order_by('username')
                for u in users:
                    ws.append([u.username, '', '', u.last_name +
                               ' ' + u.first_name, u.surname])

            else:
                raise Http404

        elif form.cleaned_data['state'] == 'participants':
            list_participants_weight = self.se.list_participants_weight()
            for e in list_participants_weight:
                u = e[0]
                ws.append([u.username, e[1], u.last_name +
                           ' ' + u.first_name, u.surname])

        elif form.cleaned_data['state'] == 'registrants':
            list_registrants_weight = self.se.list_registrants_weight()
            for e in list_registrants_weight:
                u = e[0]
                ws.append([u.username, e[1], u.last_name +
                           ' ' + u.first_name, u.surname])
        else:
            raise Http404

        # Return the file
        response = HttpResponse(save_virtual_workbook(
            wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = (
            'attachment; filename="' + self.se.__str__() + '.xlsx"')
        return response


class SharedEventUploadXlsx(GroupPermissionMixin, FormView, GroupLateralMenuMixin):
    """
    Upload Excel.
    """
    form_class = SharedEventUploadXlsxForm
    perm_codename = None  # Checked in dispatch()

    def dispatch(self, request, *args, **kwargs):
        try:
            self.group = Group.objects.get(name=kwargs['group_name'])
            self.se = SharedEvent.objects.get(pk=kwargs['pk'])
        except ObjectDoesNotExist:
            raise Http404

        # Permission
        if self.se.done is True:
            raise PermissionDenied
        if not (request.user == self.se.manager or request.user.has_perm('finances.manage_sharedevent')):
            raise PermissionDenied

        return super(SharedEventUploadXlsx, self).dispatch(request, *args, **kwargs)

    def form_valid(self, form, *args, **kwargs):
        try:
            wb = load_workbook(self.request.FILES['list_user'], read_only=True)
            sheet = wb.active
            rows = sheet.rows
            next(rows)  # Skip the first row
            data = []
        except:
            raise PermissionDenied

        if form.cleaned_data['state'] == 'participants':
            isParticipant = True
        else:
            isParticipant = False

        errors = []
        nb_empty_rows = 0
        min_col = sheet.min_column - 1
        i = 1

        # Enregistrement des pondérations
        for row in rows:
            i += 1
            try:
                if row[min_col].value and row[min_col + 1].value:
                    username = row[min_col].value.strip()  # Should be a str
                    if User.objects.filter(username=username).count() > 0:
                        user = User.objects.get(username=username)
                        try:
                            # Should be an int. Else, raise an error
                            pond = int(row[min_col+1].value)
                            if pond > 0:
                                self.se.change_weight(
                                    user, pond, isParticipant)
                        except:
                            errors.append(
                                "Erreur avec " + username + " (ligne n*" + str(i) + "). A priori pas ajouté.")
                    else:
                        errors.append("L'utilisateur " + username +
                                      " n'existe pas. (ligne n*" + str(i) + ").")
                else:
                    nb_empty_rows += 1
            except:
                errors.append("Erreur avec la ligne n*" +
                              str(i) + ". Pas ajouté.")

        errors_count = len(errors)
        error_message = ""
        if errors_count >= 1:
            error_message = str(errors_count) + \
                " erreur(s) pendant l'ajout : \n - "
            if sheet.max_row - nb_empty_rows - 1 == errors_count:
                error_message += "Aucune donnée ne peut être importée (Vérifiez le format et la syntaxe du contenu du fichier)"
            else:
                error_message += "\n - ".join(errors)
            messages.warning(self.request, error_message)
            if i - nb_empty_rows - (1 + errors_count) > 0:
                messages.success(self.request, "Les " + str(i - nb_empty_rows -
                                                            (1 + errors_count)) + " autres utilisateurs ont bien été ajoutés.")
        else:
            messages.success(self.request, "Les " + str(i-1) +
                             " utilisateurs ont bien été ajoutés.")

        return super(SharedEventUploadXlsx, self).form_valid(form)

    def get_success_url(self):
        return reverse('url_sharedevent_manage_users',
                       kwargs={'group_name': self.group.name, 'pk': self.se.pk})


@csrf_exempt
def self_lydia_callback(request):
    """
    Function to catch the callback from Lydia after a payment.

    Create all objects needed to have a proper sale in the database, and credit
    the client.

    :param GET['user_pk']: pk of the client, mandatory.
    :param POST['currency']: icon of the currency, for instance EUR, mandatory.
    :param POST['request_id']: request id from lydia, mandatory.
    :param POST['amount']: amount of the transaction, in 'currency', mandatory.
    :param POST['signed']: internal to Lydia ?
    :param POST['transaction_identifier']: transaction id from Lydia,
    mandatory.
    :param POST['vendor_token']: public key of the association, mandatory.
    :param POST['sig']: signatory generated by Lydia to identify the
    transaction, mandator.
    :type GET['user_pk']: positiv integer
    :type POST['currency']: string
    :type POST['request_id: string
    :type POST['amount']: float (decimal)
    :type POST['signed']: boolean
    :type POST['transaction_identifier']: string
    :type POST['vendor_token']: string
    :type POST['sig']: string

    :note:: Even if some parameters tend to be useless (signed, request_id),
    they are mandatory because used to generated the signatory and verify the
    transaction.

    :raises: PermissionDenied if signatory generated is not sig.
    :returns: 300 if the user_pk doesn't match an user.
    :returns: 300 if a parameter is missing.
    :returns: 200 if all's good.
    :rtype: Http request
    """
    params_dict = {
        "currency": request.POST.get("currency"),
        "request_id": request.POST.get("request_id"),
        "amount": request.POST.get("amount"),
        "signed": request.POST.get("signed"),
        "transaction_identifier": request.POST.get("transaction_identifier"),
        "vendor_token": request.POST.get("vendor_token"),
        "sig": request.POST.get("sig")
    }
    if verify_token_algo_lydia(params_dict, settings_safe_get("LYDIA_API_TOKEN").get_value()) is True:
        try:
            lydia = LydiaOnline.objects.create(
                sender=User.objects.get(pk=request.GET.get('user_pk')),
                recipient=User.objects.get(username='AE_ENSAM'),
                amount=decimal.Decimal(params_dict['amount']),
                id_from_lydia=params_dict['transaction_identifier']
            )
            recharging = Recharging.objects.create(
                sender=User.objects.get(pk=request.GET.get('user_pk')),
                operator=User.objects.get(pk=request.GET.get('user_pk')),
                payment_solution=lydia.paymentsolution_ptr
            )
            recharging.pay()
        except KeyError:
            return HttpResponse('300')
        except ObjectDoesNotExist:
            return HttpResponse('300')
        return HttpResponse('200')
    else:
        raise PermissionDenied


def verify_token_algo_lydia(params, token):
    """
    Verify request parameters according to Lydia's algorithm.

    If parameters are valid, the request is authenticated to be from Lydia and
    can be safely used.
    :note:: sig must be contained in the parameters dictionary.

    :warning:: token is private and must never be revealed.

    :param params: all parameters, including sig, mandatory.
    :type params: python dictionary
    :param token: token to be compared, mandatory.
    :type token: string

    :returns: True if parameters are valid, False else.
    :rtype: Boolean
    """
    try:
        sig = params['sig']
        del params['sig']
        h_sig_table = []
        sorted_params = sorted(params.items(), key=operator.itemgetter(0))
        for p in sorted_params:
            h_sig_table.append(p[0] + '=' + p[1])
        h_sig = '&'.join(h_sig_table)
        h_sig += '&' + token
        h_sig_hash = hashlib.md5(h_sig.encode())
        return h_sig_hash.hexdigest() == sig

    except KeyError:
        return False
